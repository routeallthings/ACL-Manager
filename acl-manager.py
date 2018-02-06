#!/usr/bin/env python

'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
INSTALL netmiko (pip install netmiko)
INSTALL textfsm (pip install textfsm)
INSTALL openpyxl (pip install openpyxl)
INSTALL fileinput (pip install fileinput)
INSTALL xlhelper (python -m pip install git+git://github.com/routeallthings/xlhelper.git)
'''

#Module Imports (Native)
import re
import getpass
import os
import unicodedata
import csv
import threading
import time
import sys

#Module Imports (Non-Native)
try:
	import netmiko
	from netmiko import ConnectHandler
except ImportError:
	netmikoinstallstatus = fullpath = raw_input ('Netmiko module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in netmikoinstallstatus.upper() or "YES" in netmikoinstallstatus.upper():
		os.system('python -m pip install netmiko')
		import netmiko
		from netmiko import ConnectHandler
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of netmiko. Please install manually and retry"
		sys.exit()
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of Pandas. Please install manually and retry'
		sys.exit()
#
try:
	import fileinput
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()
# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper		
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install git+git://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		sys.exit()
#######################################
#Functions
def GetACLs(device,acllist,aclfolder):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		print 'Successfully connected to seed device ' + devicehostname
		for acl in acllist:
			aclname = acl + '.txt'
			aclpath = aclfolder + '\\' + aclname
			if not os.path.exists(aclpath):
				print acl + ' is missing from the DB folder. Connecting to first device and downloading a current copy.'
				aclcommand = 'show run | s ip.access-list.*' + acl
				sshresult = sshnet_connect.send_command(aclcommand)
				f = open(aclpath,'w')
				f.write(sshresult)
				f.close()
			if os.path.exists(aclpath):
				aclsync = 1
				aclq = ''
				print acl + ' already exists, comparing to seed device for any changes.'
				openaclf = open(aclpath,'r')
				openacl = openaclf.readlines()
				openaclf.close()
				# CompareACLs
				aclcommand = 'show run | s ip.access-list.*' + acl
				sshresult = sshnet_connect.send_command(aclcommand)
				sshresultlistu = []
				sshresultlist = []
				openacllist = []
				sshresultlistu = sshresult.split('\n')
				for sshresultu in sshresultlistu:
					sshresult = sshresultu.encode('utf-8').lstrip()
					sshresultlist.append(sshresult)
				for openaclline in openacl:
					aclline = openaclline.strip('\n').lstrip()
					openacllist.append(aclline)
				if sshresultlist == openacllist:
					aclsync = 0
				if sshresult == '':
					print 'No ACL with the name ' + acl + ' exists on the switch.'
					aclsync = 0
				if aclsync == 1:
					if not 'y' in aclq.lower() or not 'n' in aclq.lower() or overrideacl == 1:
						aclq = acl + ' is out of sync with the seed device. Overwrite local copy with copy from seed device (Y/N)?:'
						aclq = raw_input(aclq)
					if 'y' in aclq or overrideacl == 1:
						print 'Updating ' + acl + ' with the copy from the seed device.'
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error while gathering data with ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of DB'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''

def UpdateACLs(device,acllist,aclfolder):
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	#Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		for acl in acllist:
			# Get Local DB copy of ACL
			try:
				inboundacl = acl.get('Inbound ACL').encode('utf-8')
			except:
				inboundacl = None
			try:
				outboundacl = acl.get('Outbound ACL').encode('utf-8')
			except:
				outboundacl = None
			interfacenumber = str(acl.get('VLAN #'))
			interfacename = acl.get('VLAN Name').encode('utf-8')
			# Check for interface existence
			intcheckcmd = 'show interface vlan ' + interfacenumber
			intcheck = sshnet_connect.send_command(intcheckcmd)
			if 'line' in intcheck:
				intcheckv = 1
			else:
				intcheckv = 0
			# Create the VLANs
			if intcheckv == 1:
				if not inboundacl == None:
					# Get local DB copy of ACL
					aclname = inboundacl + '.txt'
					aclpath = aclfolder + '\\' + aclname
					if os.path.exists(aclpath):
						openaclf = open(aclpath,'r')
						openacl = openaclf.readlines()
						openaclf.close()
						if openacl == '':
							print 'Loaded the ACL from the DB but it contained no data'
							openaclv = 0
						else:
							openaclv = 1
						if openaclv == 1:
							ListOfCommands = []
							# Append Remove
							removeaclextcommand = 'no ip access-list extended ' + inboundacl
							removeaclstcommand = 'no ip access-list standard ' + inboundacl
							ListOfCommands.append(removeaclextcommand)
							ListOfCommands.append(removeaclstcommand)
							# Append ACL
							for aclline in openacl:
								aclline = aclline.strip('\n')
								ListOfCommands.append(aclline)
							# Append Interface Commands
							interfacecommand = 'interface vlan ' + interfacenumber
							ListOfCommands.append(interfacecommand)
							addtointerfacecommand = 'ip access-group ' + inboundacl + ' in'
							ListOfCommands.append(addtointerfacecommand)
							# Apply Changes
							FullOutput = sshnet_connect.send_config_set(ListOfCommands)
					else:
						print 'Error with getting local DB copy for ' + interfacename + '. Skipping...'
				if not outboundacl == None:
					# Get local DB copy of ACL
					aclname = outboundacl + '.txt'
					aclpath = aclfolder + '\\' + aclname
					if os.path.exists(aclpath):
						openaclf = open(aclpath,'r')
						openacl = openaclf.readlines()
						openaclf.close()
						if openacl == '':
							print 'Loaded the ACL from the DB but it contained no data'
							openaclv = 0
						else:
							openaclv = 1
						if openaclv == 1:
							ListOfCommands = []
							# Append Remove
							removeaclextcommand = 'no ip access-list extended ' + outboundacl
							removeaclstcommand = 'no ip access-list standard ' + outboundacl
							ListOfCommands.append(removeaclextcommand)
							ListOfCommands.append(removeaclstcommand)
							# Append ACL
							for aclline in openacl:
								aclline = aclline.strip('\n')
								ListOfCommands.append(aclline)
							# Append Interface Commands
							interfacecommand = 'interface vlan ' + interfacenumber
							ListOfCommands.append(interfacecommand)
							addtointerfacecommand = 'ip access-group ' + outboundacl + ' out'
							ListOfCommands.append(addtointerfacecommand)
							# Apply Changes
							FullOutput = sshnet_connect.send_config_set(ListOfCommands)
					else:
						print 'Error with getting local DB copy for ' + interfacename + '. Skipping...'
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error with sending commands to ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of switches'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
#########################################
print ''
print 'ACL Manager'
print '############################################################'
print 'The purpose of this tool is to use a XLSX import to control'
print 'and set ACLs on various interfaces that are assigned.'
print 'Please fill in the config tab on the templated XLSX'
print 'sheet, along with all the data that you want to test.'
print '############################################################'
print ''
print '----Questions that need answering----'
excelfilelocation = raw_input('File to load the excel data from (e.g. C:/Python27/acl-datatemplate.xlsx):')
if excelfilelocation == '':
	excelfilelocation = 'C:/Python27/acl-datatemplate.xlsx'
excelfilelocation = excelfilelocation.replace('"', '')
# Load Configuration Variables
configdict = {}
for configvariables in xlhelper.sheet_to_dict(excelfilelocation,'Config'):
	try:
		configvar = configvariables.get('Variable').encode('utf-8')
		configval = configvariables.get('Value').encode('utf-8')
	except:
		configvar = configvariables.get('Variable')
		configval = configvariables.get('Value')
	configdict[configvar] = configval
# Username Variables/Questions
sshusername = configdict.get('Username')
if 'NA' == sshusername:
	sshusername = raw_input('What is the username you will use to login to the devices?:')
sshpassword = configdict.get('Password')
if 'NA' == sshpassword:
	sshpassword = getpass.getpass('What is the password you will use to login to the devices?:')
enablesecret = configdict.get('EnableSecret')
if 'NA' == enablesecret:
	enablesecret = getpass.getpass('What is the enable password you will use to access the devices?:')
# Rest of the Config Variables
databaselocation = configdict.get('DatabaseFolder')
if databaselocation == None:
	databaselocation = r'C:\ACLManager\DB'
seeddeviceip = configdict.get('SeedDeviceIP')
if seeddeviceip == None:
	seeddeviceip = raw_input('Need the IP of a seed device to pull ACLs from (e.g. 10.1.1.1):')
overrideacl = configdict.get('OverrideACLs')
if overrideacl == None:
	overrideacl = 0
updatelocaldb = raw_input('Do you want to update your local DB (Y/N)?:')
# Get Other Libraries
devicelist = []
vlanlist = []
for devices in xlhelper.sheet_to_dict(excelfilelocation,'Device IPs'):
	devicelist.append(devices)
for vlans in xlhelper.sheet_to_dict(excelfilelocation,'VLAN List'):
	vlanlist.append(vlans)
# Create Database folder if its missing
newinstall = 0
if not os.path.exists(databaselocation):
	os.makedirs(databaselocation)
	newinstall = 1
#### ACL DB Check (Up to date)
print 'Starting Database Check'
for device in devicelist:
	deviceip = device.get('IP').encode('utf-8')
	if deviceip == seeddeviceip:
		seeddevice = device
fullacllist = []
for vlans in vlanlist:
	try:
		inboundacl = vlans.get('Inbound ACL').encode('utf-8')
	except:
		inboundacl = None
	try:
		outboundacl = vlans.get('Outbound ACL').encode('utf-8')
	except:
		outboundacl = None
	# Check for Duplicates
	duplicatevlanacl = 0
	for acls in fullacllist:
		if inboundacl == acls:
			duplicatevlanacl = 1
		if outboundacl == acls:
			duplicatevlanacl = 1		
	# Add to VLAN list
	if not inboundacl == None and not duplicatevlanacl == 1:
		fullacllist.append(inboundacl)
	if not outboundacl == None and not duplicatevlanacl == 1:
		fullacllist.append(outboundacl)
if 'y' in updatelocaldb.lower():
	GetACLs(seeddevice,fullacllist,databaselocation)
	print 'Completed update of local DB'
else:
	print 'Skipping Database Check because of user input'
#### Update devices with new ACLs
continueq = raw_input('Do you want to continue with updating all devices from local copy (Y/N)?:')
if not 'y' in continueq.lower() and not 'n' in continueq.lower():
	continueq = raw_input('Please enter yes or no to the above question (Y/N)?:')
if 'n' in continueq.lower():
	print 'Exiting script...'
	sys.exit()
if __name__ == "__main__":
	# Start Threads
	if 'y' in continueq.lower():
		print 'Starting update on all switches'
		for device in devicelist:	
			deviceip = device.get('IP').encode('utf-8')
			print "Spawning Thread for " + deviceip
			t = threading.Thread(target=UpdateACLs, args=(device,vlanlist,databaselocation))
			t.start()
		main_thread = threading.currentThread()
		# Join All Threads
		for it_thread in threading.enumerate():
			if it_thread != main_thread:
				it_thread.join()
print 'ACL Manager has completed updating the switches. Exiting..'
